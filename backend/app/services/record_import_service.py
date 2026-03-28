"""Services for admin-only XLSX record imports."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models import (
    Author,
    AuthorType,
    KeywordLocation,
    KeywordName,
    KeywordRecord,
    Language,
    Lettering,
    LoanType,
    PublicationType,
    Publisher,
    Record,
    RecordAuthor,
    RecordCondition,
    Restriction,
    WorkStatus,
)
from app.utils.phonetics import generate_phonetic_codes


HEADER_MAP = {
    "titel": "title",
    "signaturneu": "signature",
    "signatur2": "signature2",
    "signatur": "signature",
    "publiknr": "bibl_nr",
    "publikation": "publicationtype",
    "jahr": "year",
    "isbniss": "isbn",
    "seitenzahl": "number_pages",
    "auflage": "edition",
    "reihe": "reihe",
    "band": "volume",
    "jahrgang": "jahrgang",
    "eingabedat": "enter_information",
    "zustand": "record_condition",
    "schlagworter": "keyword_record",
    "schlagwoerter": "keyword_record",
    "indexe": "indecies",
    "entleihbar": "loantype",
    "schrift": "lettering",
    "sprache": "language",
    "eingabedat2": "enter_date",
    "aussonderdat": "sort_out_date",
    "bemerkung": "comment",
    "autor": "author",
    "orte": "keyword_location",
    "familiennamen": "keyword_name",
    "verlag": "publisher",
    "verlagsort": "publisher_town",
}

KNOWN_TITLES = ["Prof. Dr.", "Dr. jur.", "Prof.", "Dr."]


@dataclass
class ImportResult:
    imported: int
    skipped: int
    errors: List[Dict[str, Any]]


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    normalized = str(value).strip().lower()
    normalized = normalized.replace("ä", "a").replace("ö", "o").replace("ü", "u").replace("ß", "ss")
    return re.sub(r"[^a-z0-9]", "", normalized)


def _to_string(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _split_values(value: Optional[str]) -> List[str]:
    if not value:
        return []
    return [item.strip() for item in re.split(r"[;,]", value) if item and item.strip()]


def _split_comma_values(value: Optional[str]) -> List[str]:
    if not value:
        return []
    return [item.strip() for item in value.split(",") if item and item.strip()]


def _remove_excel_linebreak_artifact(value: Optional[str], row_index: int, field_name: str, errors: List[Dict[str, Any]]) -> Optional[str]:
    if not value:
        return value

    if "_x000d_" in value:
        cleaned = value.replace("_x000d_", "\n")
        errors.append(
            {
                "row": row_index,
                "message": f"Replaced _x000d_ with line break in field '{field_name}'",
            }
        )
        return cleaned if cleaned else None

    return value


def _extract_authortype(value: str) -> tuple[str, Optional[str]]:
    """Extract authortype in (...) or [...] and return cleaned text + authortype."""
    match = re.search(r"(\([^)]*\)|\[[^\]]*\])", value)
    if not match:
        return value.strip(), None

    authortype = match.group(0).strip()
    cleaned = (value[: match.start()] + value[match.end() :]).strip()
    return cleaned, authortype


def _split_author_entries(value: Optional[str]) -> List[str]:
    if not value:
        return []
    return [item.strip() for item in value.split(";") if item and item.strip()]


def parse_excel_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    raise ValueError(f"Invalid Excel date value: {value}")


def _parse_excel_date_with_log(value: Any, row_index: int, field_name: str, errors: List[Dict[str, Any]]) -> Optional[date]:
    try:
        return parse_excel_date(value)
    except ValueError:
        errors.append(
            {
                "row": row_index,
                "message": f"Invalid date in field '{field_name}' cleared",
            }
        )
        return None


def parse_author_entries(raw_value: Optional[str]) -> List[Dict[str, Optional[str]]]:
    """Parse author string into structured author entries preserving order."""
    entries: List[Dict[str, Optional[str]]] = []
    for author_entry in _split_author_entries(raw_value):
        parts = [part.strip() for part in author_entry.split(",") if part and part.strip()]
        if not parts:
            continue

        last_name = parts[0]
        remainder = ",".join(parts[1:]).strip() if len(parts) > 1 else ""

        # Authortype may be provided in either round or square brackets.
        if remainder:
            remainder, authortype = _extract_authortype(remainder)
        else:
            last_name, authortype = _extract_authortype(last_name)

        title = None
        first_name = remainder
        for candidate in KNOWN_TITLES:
            if first_name.startswith(candidate):
                title = candidate
                first_name = first_name[len(candidate):].strip()
                break

        entries.append(
            {
                "last_name": last_name,
                "first_name": first_name or None,
                "title": title,
                "authortype": authortype,
            }
        )

    return entries


def _get_or_create_case_insensitive(
    db: Session,
    model,
    field_name: str,
    value: Optional[str],
    create_kwargs: Optional[Dict[str, Any]] = None,
):
    if not value:
        return None
    cleaned = value.strip()
    if not cleaned:
        return None

    field = getattr(model, field_name)
    existing = db.query(model).filter(func.lower(field) == cleaned.lower()).first()
    if existing:
        return existing

    kwargs = {field_name: cleaned}
    if create_kwargs:
        kwargs.update(create_kwargs)
    obj = model(**kwargs)
    db.add(obj)
    db.flush()
    return obj


def _get_or_create_defaults(db: Session):
    restriction = db.query(Restriction).filter(func.lower(Restriction.name) == "none").first()
    if not restriction:
        restriction = Restriction(name="none")
        db.add(restriction)
        db.flush()

    workstatus = db.query(WorkStatus).filter(func.lower(WorkStatus.status) == "not yet").first()
    if not workstatus:
        raise ValueError("Default WorkStatus 'not yet' is missing")

    blank_authortype = db.query(AuthorType).filter(func.lower(AuthorType.authortype) == "blank").first()
    if not blank_authortype:
        blank_authortype = AuthorType(authortype="Blank")
        db.add(blank_authortype)
        db.flush()

    return restriction, workstatus, blank_authortype


def import_records_from_xlsx(file_bytes: bytes, db: Session, current_user_id) -> ImportResult:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)

    if len(workbook.worksheets) != 1:
        raise ValueError("Excel file must contain exactly one worksheet")

    sheet = workbook.worksheets[0]
    header_cells = [cell.value for cell in sheet[1]]
    headers: Dict[int, str] = {}
    for index, header in enumerate(header_cells):
        normalized = _normalize_header(header)
        if normalized in HEADER_MAP:
            headers[index] = HEADER_MAP[normalized]

    required_source_fields = {"signature", "bibl_nr"}
    if not required_source_fields.issubset(set(headers.values())):
        missing = sorted(required_source_fields - set(headers.values()))
        raise ValueError(f"Missing required columns in header row: {', '.join(missing)}")

    restriction, workstatus, blank_authortype = _get_or_create_defaults(db)

    imported = 0
    skipped = 0
    errors: List[Dict[str, Any]] = []

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value in (None, "") for value in row):
            continue

        row_data: Dict[str, Any] = {}
        for col_index, mapped_field in headers.items():
            if col_index < len(row):
                row_data[mapped_field] = row[col_index]

        signature = _remove_excel_linebreak_artifact(_to_string(row_data.get("signature")), row_index, "signature", errors)
        signature2 = _remove_excel_linebreak_artifact(_to_string(row_data.get("signature2")), row_index, "signature2", errors)
        bibl_nr = _remove_excel_linebreak_artifact(_to_string(row_data.get("bibl_nr")), row_index, "bibl_nr", errors)

        if not signature or not bibl_nr:
            skipped += 1
            errors.append(
                {
                    "row": row_index,
                    "message": "Missing required values for Signatur Neu or PublikNr",
                }
            )
            continue

        savepoint = db.begin_nested()
        try:
            loantype_raw = _remove_excel_linebreak_artifact(
                _to_string(row_data.get("loantype")),
                row_index,
                "loantype",
                errors,
            )
            loan_value = None
            subtype_value = None
            if loantype_raw:
                loan_parts = [part.strip() for part in loantype_raw.split(";", 1)]
                loan_value = loan_parts[0] if loan_parts and loan_parts[0] else None
                subtype_value = loan_parts[1] if len(loan_parts) > 1 and loan_parts[1] else None

            publicationtype = _get_or_create_case_insensitive(
                db,
                PublicationType,
                "publicationtype",
                _remove_excel_linebreak_artifact(
                    _to_string(row_data.get("publicationtype")),
                    row_index,
                    "publicationtype",
                    errors,
                ),
            )
            record_condition = _get_or_create_case_insensitive(
                db,
                RecordCondition,
                "condition",
                _remove_excel_linebreak_artifact(
                    _to_string(row_data.get("record_condition")),
                    row_index,
                    "record_condition",
                    errors,
                ),
            )
            loantype = None
            if loan_value:
                loantype = (
                    db.query(LoanType)
                    .filter(func.lower(LoanType.loan) == loan_value.lower())
                    .filter(
                        func.lower(LoanType.subtype) == subtype_value.lower()
                        if subtype_value
                        else LoanType.subtype.is_(None)
                    )
                    .first()
                )
                if not loantype:
                    loantype = LoanType(loan=loan_value, subtype=subtype_value)
                    db.add(loantype)
                    db.flush()

            lettering = _get_or_create_case_insensitive(
                db,
                Lettering,
                "lettering",
                _remove_excel_linebreak_artifact(
                    _to_string(row_data.get("lettering")),
                    row_index,
                    "lettering",
                    errors,
                ),
            )
            publisher_name = _remove_excel_linebreak_artifact(
                _to_string(row_data.get("publisher")),
                row_index,
                "publisher",
                errors,
            )
            publisher = _get_or_create_case_insensitive(
                db,
                Publisher,
                "companyname",
                publisher_name,
                create_kwargs={
                    "town": _remove_excel_linebreak_artifact(
                        _to_string(row_data.get("publisher_town")),
                        row_index,
                        "publisher_town",
                        errors,
                    )
                },
            )

            title = _remove_excel_linebreak_artifact(
                _to_string(row_data.get("title")),
                row_index,
                "title",
                errors,
            ) or signature
            record = Record(
                title=title,
                signature=signature,
                signature2=signature2,
                year=_remove_excel_linebreak_artifact(
                    _to_string(row_data.get("year")),
                    row_index,
                    "year",
                    errors,
                ),
                isbn=_remove_excel_linebreak_artifact(
                    _to_string(row_data.get("isbn")),
                    row_index,
                    "isbn",
                    errors,
                ),
                number_pages=_remove_excel_linebreak_artifact(
                    _to_string(row_data.get("number_pages")),
                    row_index,
                    "number_pages",
                    errors,
                ),
                edition=_remove_excel_linebreak_artifact(
                    _to_string(row_data.get("edition")),
                    row_index,
                    "edition",
                    errors,
                ),
                reihe=_remove_excel_linebreak_artifact(
                    _to_string(row_data.get("reihe")),
                    row_index,
                    "reihe",
                    errors,
                ),
                volume=_remove_excel_linebreak_artifact(
                    _to_string(row_data.get("volume")),
                    row_index,
                    "volume",
                    errors,
                ),
                jahrgang=_remove_excel_linebreak_artifact(
                    _to_string(row_data.get("jahrgang")),
                    row_index,
                    "jahrgang",
                    errors,
                ),
                bibl_nr=bibl_nr,
                enter_information=_remove_excel_linebreak_artifact(
                    _to_string(row_data.get("enter_information")),
                    row_index,
                    "enter_information",
                    errors,
                ),
                indecies=_remove_excel_linebreak_artifact(
                    _to_string(row_data.get("indecies")),
                    row_index,
                    "indecies",
                    errors,
                ),
                comment=_remove_excel_linebreak_artifact(
                    _to_string(row_data.get("comment")),
                    row_index,
                    "comment",
                    errors,
                ),
                enter_date=_parse_excel_date_with_log(row_data.get("enter_date"), row_index, "enter_date", errors),
                sort_out_date=_parse_excel_date_with_log(row_data.get("sort_out_date"), row_index, "sort_out_date", errors),
                restriction_id=restriction.id,
                workstatus_id=workstatus.id,
                publicationtype_id=publicationtype.id if publicationtype else None,
                record_condition_id=record_condition.id if record_condition else None,
                loantype_id=loantype.id if loantype else None,
                lettering_id=lettering.id if lettering else None,
                publisher_id=publisher.id if publisher else None,
                created_by=current_user_id,
            )
            db.add(record)
            db.flush()

            for keyword_name in _split_values(_to_string(row_data.get("keyword_record"))):
                keyword = _get_or_create_case_insensitive(db, KeywordRecord, "name", keyword_name)
                if keyword and not keyword.c_search:
                    c_search, dblmeta_1, dblmeta_2 = generate_phonetic_codes(keyword.name)
                    keyword.c_search = c_search
                    keyword.dblmeta_1 = dblmeta_1
                    keyword.dblmeta_2 = dblmeta_2
                if keyword and keyword not in record.keywords_records:
                    record.keywords_records.append(keyword)

            for keyword_location_name in _split_comma_values(
                _remove_excel_linebreak_artifact(
                    _to_string(row_data.get("keyword_location")),
                    row_index,
                    "keyword_location",
                    errors,
                )
            ):
                keyword_location = _get_or_create_case_insensitive(db, KeywordLocation, "name", keyword_location_name)
                if keyword_location and not keyword_location.c_search:
                    c_search, dblmeta_1, dblmeta_2 = generate_phonetic_codes(keyword_location.name)
                    keyword_location.c_search = c_search
                    keyword_location.dblmeta_1 = dblmeta_1
                    keyword_location.dblmeta_2 = dblmeta_2
                if keyword_location and keyword_location not in record.keywords_locations:
                    record.keywords_locations.append(keyword_location)

            for keyword_name_value in _split_comma_values(
                _remove_excel_linebreak_artifact(
                    _to_string(row_data.get("keyword_name")),
                    row_index,
                    "keyword_name",
                    errors,
                )
            ):
                keyword_name = _get_or_create_case_insensitive(db, KeywordName, "name", keyword_name_value)
                if keyword_name and not keyword_name.c_search:
                    c_search, dblmeta_1, dblmeta_2 = generate_phonetic_codes(keyword_name.name)
                    keyword_name.c_search = c_search
                    keyword_name.dblmeta_1 = dblmeta_1
                    keyword_name.dblmeta_2 = dblmeta_2
                if keyword_name and keyword_name not in record.keywords_names:
                    record.keywords_names.append(keyword_name)

            for language_name in _split_values(_to_string(row_data.get("language"))):
                language = _get_or_create_case_insensitive(db, Language, "language", language_name)
                if language and language not in record.languages:
                    record.languages.append(language)

            author_entries = parse_author_entries(
                _remove_excel_linebreak_artifact(
                    _to_string(row_data.get("author")),
                    row_index,
                    "author",
                    errors,
                )
            )
            for position, author_entry in enumerate(author_entries, start=1):
                author = _get_or_create_case_insensitive(
                    db,
                    Author,
                    "last_name",
                    author_entry["last_name"],
                    create_kwargs={
                        "first_name": author_entry["first_name"],
                        "title": author_entry["title"],
                        "created_by": current_user_id,
                    },
                )

                if author and author_entry["first_name"] and not author.first_name:
                    author.first_name = author_entry["first_name"]
                if author and author_entry["title"] and not author.title:
                    author.title = author_entry["title"]

                authortype_value = author_entry["authortype"]
                authortype = blank_authortype
                if authortype_value:
                    authortype = _get_or_create_case_insensitive(
                        db,
                        AuthorType,
                        "authortype",
                        authortype_value,
                    )

                if author:
                    relation = RecordAuthor(
                        record_id=record.id,
                        author_id=author.id,
                        authortype_id=authortype.id if authortype else blank_authortype.id,
                        order=position,
                        created_by=current_user_id,
                    )
                    db.add(relation)

            savepoint.commit()
            imported += 1
        except Exception as exc:
            savepoint.rollback()
            skipped += 1
            errors.append({"row": row_index, "message": str(exc)})

    db.commit()
    return ImportResult(imported=imported, skipped=skipped, errors=errors)
