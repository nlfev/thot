from tests.conftest import auth_headers_and_csrf
"""Tests for admin XLSX record import routes."""

from io import BytesIO

from openpyxl import Workbook

from app.models import (
    Author,
    AuthorType,
    KeywordLocation,
    KeywordName,
    KeywordRecord,
    Language,
    LoanType,
    PublicationType,
    Publisher,
    Record,
    RecordAuthor,
    Role,
    User,
    UserRole,
    WorkStatus,
    WorkStatusArea,
)
from app.utils import create_access_token, hash_password


def _create_user_with_role(db, username: str, role_name: str) -> User:
    role = db.query(Role).filter(Role.name == role_name).first()
    if role is None:
        role = Role(name=role_name, description=f"{role_name} role")
        db.add(role)
        db.flush()

    user = User(
        username=username,
        email=f"{username}@example.com",
        hashed_password=hash_password("TestPassword123!"),
        active=True,
    )
    db.add(user)
    db.flush()

    db.add(UserRole(user_id=user.id, role_id=role.id, active=True))
    db.commit()
    db.refresh(user)
    return user



# Neue Hilfsfunktion: Auth-Header + CSRF für Tests
def _auth_headers_and_csrf(user: User):
    return auth_headers_and_csrf(user)


def _ensure_default_workstatus(db):
    area = db.query(WorkStatusArea).filter(WorkStatusArea.area == "record").first()
    if area is None:
        area = WorkStatusArea(area="record")
        db.add(area)
        db.flush()

    default_status = db.query(WorkStatus).filter(WorkStatus.status == "not yet").first()
    if default_status is None:
        default_status = WorkStatus(status="not yet", workstatus_area_id=area.id)
        db.add(default_status)
        db.flush()

    db.commit()


def _build_xlsx_bytes(rows, include_second_sheet: bool = False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Import"
    sheet.append(
        [
            "Titel",
            "Signatur Neu",
            "Signatur2",
            "PublikNr",
            "Publikation",
            "Schlagwörter",
            "Orte",
            "Familiennamen",
            "Entleihbar",
            "Jahr",
            "ISBN_ISS",
            "Seitenzahl",
            "Auflage",
            "Reihe",
            "Band",
            "Jahrgang",
            "Sprache",
            "Autor",
            "Eingabedat2",
        ]
    )

    for row in rows:
        sheet.append(row)

    if include_second_sheet:
        workbook.create_sheet("Second")

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_record_import_admin_only(client, db):
    _ensure_default_workstatus(db)
    support_user = _create_user_with_role(db, "import_support", "support")

    file_bytes = _build_xlsx_bytes([
        ["Title 1", "SIG-1", "", "BIB-1", "Book", "Alpha", "", "", "", "", "", "", "", "", "", "", "Deutsch", "Miller, Max", "2025-01-10"]
    ])

    headers, cookies = _auth_headers_and_csrf(support_user)
    client.cookies.clear()
    for k, v in cookies.items():
        client.cookies.set(k, v)
    response = client.post(
        "/api/v1/admin/records-import/xlsx",
        headers=headers,
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 403


def test_record_import_creates_duplicate_signature_and_metadata(client, db):
    _ensure_default_workstatus(db)
    admin_user = _create_user_with_role(db, "import_admin", "admin")

    file_bytes = _build_xlsx_bytes(
        [
            ["Title 1", "SIG-1", "", "BIB-1", "Book", "Alpha", "", "", "", "", "", "", "", "", "", "", "Deutsch", "Miller, Max", "2025-01-10"],
            ["Title 2", "SIG-1", "", "BIB-2", "Book", "Alpha", "", "", "", "", "", "", "", "", "", "", "Deutsch", "Miller, Max", "2025-01-11"],
        ]
    )

    headers, cookies = _auth_headers_and_csrf(admin_user)
    client.cookies.clear()
    for k, v in cookies.items():
        client.cookies.set(k, v)
    response = client.post(
        "/api/v1/admin/records-import/xlsx",
        headers=headers,
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["imported"] == 2
    assert payload["skipped"] == 0
    assert payload["errors"] == []

    assert db.query(Record).filter(Record.signature == "SIG-1").count() == 2
    assert db.query(PublicationType).filter(PublicationType.publicationtype == "Book").first() is not None
    assert db.query(Language).filter(Language.language == "Deutsch").first() is not None
    keyword = db.query(KeywordRecord).filter(KeywordRecord.name == "Alpha").first()
    assert keyword is not None
    assert keyword.c_search is not None


def test_record_import_returns_error_report_for_missing_required_values(client, db):
    _ensure_default_workstatus(db)
    admin_user = _create_user_with_role(db, "import_admin_errors", "admin")

    file_bytes = _build_xlsx_bytes(
        [
            ["Title 1", None, "", "BIB-1", "Book", "Alpha", "", "", "", "", "", "", "", "", "", "", "Deutsch", "Miller, Max", "2025-01-10"],
            ["Title 2", "SIG-2", "", None, "Book", "Alpha", "", "", "", "", "", "", "", "", "", "", "Deutsch", "Miller, Max", "2025-01-10"],
        ]
    )

    headers, cookies = _auth_headers_and_csrf(admin_user)
    client.cookies.clear()
    for k, v in cookies.items():
        client.cookies.set(k, v)
    response = client.post(
        "/api/v1/admin/records-import/xlsx",
        headers=headers,
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["imported"] == 0
    assert payload["skipped"] == 2
    assert len(payload["errors"]) == 2
    assert payload["errors"][0]["row"] == 2
    assert payload["errors"][1]["row"] == 3


def test_record_import_requires_single_sheet(client, db):
    _ensure_default_workstatus(db)
    admin_user = _create_user_with_role(db, "import_admin_sheets", "admin")

    file_bytes = _build_xlsx_bytes(
        [["Title 1", "SIG-1", "", "BIB-1", "Book", "Alpha", "", "", "", "", "", "", "", "", "", "", "Deutsch", "Miller, Max", "2025-01-10"]],
        include_second_sheet=True,
    )

    headers, cookies = _auth_headers_and_csrf(admin_user)
    client.cookies.clear()
    for k, v in cookies.items():
        client.cookies.set(k, v)
    response = client.post(
        "/api/v1/admin/records-import/xlsx",
        headers=headers,
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert "exactly one worksheet" in response.json()["detail"]


def test_record_import_maps_signature2_loantype_keywords_and_cleans_x000d(client, db):
    _ensure_default_workstatus(db)
    admin_user = _create_user_with_role(db, "import_admin_enhanced", "admin")

    file_bytes = _build_xlsx_bytes(
        [
            [
                "Title 1",
                "SIG-1",
                "SIG2-1_x000d_",
                "BIB-1",
                "Book",
                "Alpha",
                "Berlin, Hamburg",
                "Meyer, Schmidt",
                "Ausleihbar;Präsenz",
                "1888",
                "978-1-23-456789-0",
                "321",
                "2",
                "Sammlung X",
                "Band 4",
                "Jg. 12",
                "Deutsch",
                "Meyer, Prof. Dr. Anna [Hg.]",
                "2025-01-10",
            ]
        ]
    )

    headers, cookies = _auth_headers_and_csrf(admin_user)
    client.cookies.clear()
    for k, v in cookies.items():
        client.cookies.set(k, v)
    response = client.post(
        "/api/v1/admin/records-import/xlsx",
        headers=headers,
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["imported"] == 1
    assert any("_x000d_" in item["message"] for item in payload["errors"])

    record = db.query(Record).filter(Record.signature == "SIG-1").first()
    assert record is not None
    assert record.signature2 == "SIG2-1\n"
    assert record.year == "1888"
    assert record.isbn == "978-1-23-456789-0"
    assert record.number_pages == "321"
    assert record.edition == "2"
    assert record.reihe == "Sammlung X"
    assert record.volume == "Band 4"
    assert record.jahrgang == "Jg. 12"

    loantype = db.query(LoanType).filter(LoanType.loan == "Ausleihbar").first()
    assert loantype is not None
    assert loantype.subtype == "Präsenz"

    berlin = db.query(KeywordLocation).filter(KeywordLocation.name == "Berlin").first()
    hamburg = db.query(KeywordLocation).filter(KeywordLocation.name == "Hamburg").first()
    assert berlin is not None and berlin.c_search is not None
    assert hamburg is not None and hamburg.c_search is not None

    meyer_keyword = db.query(KeywordName).filter(KeywordName.name == "Meyer").first()
    schmidt_keyword = db.query(KeywordName).filter(KeywordName.name == "Schmidt").first()
    assert meyer_keyword is not None and meyer_keyword.c_search is not None
    assert schmidt_keyword is not None and schmidt_keyword.c_search is not None

    author = db.query(Author).filter(Author.last_name == "Meyer").first()
    assert author is not None
    assert author.title == "Prof. Dr."
    assert author.first_name == "Anna"

    authortype = db.query(AuthorType).filter(AuthorType.authortype == "[Hg.]").first()
    assert authortype is not None
    relation = db.query(RecordAuthor).filter(RecordAuthor.record_id == record.id, RecordAuthor.author_id == author.id).first()
    assert relation is not None
    assert relation.authortype_id == authortype.id


def test_record_import_invalid_date_is_cleared_and_logged(client, db):
    _ensure_default_workstatus(db)
    admin_user = _create_user_with_role(db, "import_admin_bad_date", "admin")

    file_bytes = _build_xlsx_bytes(
        [
            [
                "Title 1",
                "SIG-DATE-1",
                "",
                "BIB-DATE-1",
                "Book",
                "Alpha",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "Deutsch",
                "Miller, Max",
                "31-31-2025",
            ]
        ]
    )

    headers, cookies = _auth_headers_and_csrf(admin_user)
    client.cookies.clear()
    for k, v in cookies.items():
        client.cookies.set(k, v)
    response = client.post(
        "/api/v1/admin/records-import/xlsx",
        headers=headers,
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["imported"] == 1
    assert payload["skipped"] == 0
    assert any("Invalid date" in item["message"] for item in payload["errors"])

    record = db.query(Record).filter(Record.signature == "SIG-DATE-1").first()
    assert record is not None
    assert record.enter_date is None


def test_record_import_stores_publisher_id(client, db):
    """Importing a row with a Verlag value must set publisher_id on the record."""
    _ensure_default_workstatus(db)
    admin_user = _create_user_with_role(db, "import_admin_publisher", "admin")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Import"
    sheet.append(
        [
            "Titel",
            "Signatur Neu",
            "Signatur2",
            "PublikNr",
            "Publikation",
            "Schlagwörter",
            "Orte",
            "Familiennamen",
            "Entleihbar",
            "Jahr",
            "ISBN_ISS",
            "Seitenzahl",
            "Auflage",
            "Reihe",
            "Band",
            "Jahrgang",
            "Sprache",
            "Autor",
            "Eingabedat2",
            "Verlag",
            "Verlagsort",
        ]
    )
    sheet.append(
        [
            "Testatlogie",
            "SIG-PUB-1",
            "",
            "BIB-PUB-1",
            "Book",
            "",
            "",
            "",
            "",
            "2024",
            "",
            "",
            "",
            "",
            "",
            "",
            "Deutsch",
            "Muster, Max",
            "2024-01-01",
            "Testverlag GmbH",
            "Berlin",
        ]
    )
    buffer = BytesIO()
    workbook.save(buffer)
    file_bytes = buffer.getvalue()

    headers, cookies = _auth_headers_and_csrf(admin_user)
    client.cookies.clear()
    for k, v in cookies.items():
        client.cookies.set(k, v)
    response = client.post(
        "/api/v1/admin/records-import/xlsx",
        headers=headers,
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["imported"] == 1, payload
    assert payload["skipped"] == 0

    publisher = db.query(Publisher).filter(Publisher.companyname == "Testverlag GmbH").first()
    assert publisher is not None, "Publisher was not created"
    assert publisher.town == "Berlin"

    record = db.query(Record).filter(Record.signature == "SIG-PUB-1").first()
    assert record is not None
    assert record.publisher_id == publisher.id, (
        f"Expected publisher_id={publisher.id} but got {record.publisher_id}"
    )
